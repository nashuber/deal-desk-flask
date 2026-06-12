"""Port of src/lib/excelParser.ts.

Reads an .xlsx workbook, auto-detects the deals + stores sheets via header
alias scoring, and returns two normalized DataFrames.
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from .data_model import normalize_deals, normalize_stores

DEAL_ALIASES: Dict[str, List[str]] = {
    "date": ["date", "month", "accountingdate", "orderdate", "period"],
    "merchant_type": ["merchanttype", "type", "vertical"],
    "merchant_segment": ["merchantsegment", "segment"],
    "grouped_parent_name": ["groupedparentname", "parentname", "merchantname", "brand", "account"],
    "territory": ["territory", "market", "region", "city"],
    "Vertical": ["vertical", "verticalname"],
    "country": ["country", "countryname"],
    "trips": ["trips", "orders", "completedtrips", "completedorders"],
    "basket": ["basket", "totalsales", "sales", "gmv", "subtotal"],
    "booking_fees": ["bookingfees", "deliveryfee", "deliveryfees", "bookingfee"],
    "service_fees": ["servicefees", "servicefee"],
    "other_uber_fees": ["otheruberfees", "markup", "otherfees", "uberfees"],
    "gross_bookings": ["grossbookings", "grossbooking", "gb"],
    "courier_payment": ["courierpayment", "courierpayments", "courierpay", "cpt"],
    "resto_payments": ["restopayments", "merchantpayments", "restaurantpayments", "restopayment"],
    "tf_disbursed": ["tfdisbursed", "promotions", "tfdisbursements"],
    "EuP": ["eup", "eaterpromotions", "eaterpromo"],
    "pass_net": ["passnet", "uberone", "uberonediscount", "passnetdiscount"],
    "ads_rev": ["adsrev", "adrevenue", "adsrevenue", "advertising"],
    "other_rev": ["otherrev", "otherrevenue"],
    "netr": ["netr", "netrevenue", "net"],
    "trip_insurance": ["tripinsurance", "insurance"],
    "support": ["support", "supportcost"],
    "money": ["money", "moneycost", "paymentsmoney"],
    "tech": ["tech", "techcost"],
    "other_variable": ["othervariable", "othervariablecost", "otherothervar"],
}

STORE_ALIASES: Dict[str, List[str]] = {
    "accounting_date": ["accountingdate", "date", "month", "period"],
    "merchant_type": ["merchanttype", "type"],
    "merchant_segment": ["merchantsegment", "segment"],
    "territory": ["territory", "market", "region"],
    "active_stores": ["activestores", "stores", "storecount", "activestorecount"],
    "store_level": ["storelevel", "level", "grain"],
    "Vertical": ["vertical", "verticalname"],
    "Region": ["region", "macroregion"],
    "country": ["country", "countryname"],
}

DEAL_DATE_FIELDS = {"date"}
STORE_DATE_FIELDS = {"accounting_date"}
DEAL_NUMERIC = {
    "trips", "basket", "booking_fees", "service_fees", "other_uber_fees",
    "gross_bookings", "courier_payment", "resto_payments", "tf_disbursed",
    "EuP", "pass_net", "ads_rev", "other_rev", "netr",
    "trip_insurance", "support", "money", "tech", "other_variable",
}
STORE_NUMERIC = {"active_stores"}

_NUM_RE = re.compile(r"[\$,\s%]")


def _norm(s: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _to_num(v: object) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return 0.0
    s = str(v).strip()
    if not s:
        return 0.0
    s = _NUM_RE.sub("", s)
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def _to_date(v: object) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    try:
        return datetime.fromisoformat(s[:10]).date().isoformat()
    except ValueError:
        pass
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%b %Y", "%B %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s


def _build_column_map(headers: List[str], aliases: Dict[str, List[str]]) -> Dict[str, int]:
    norm_headers = [_norm(h) for h in headers]
    out: Dict[str, int] = {}
    for field, cands in aliases.items():
        idx = next((i for i, h in enumerate(norm_headers) if h in cands), -1)
        if idx == -1:
            idx = next(
                (i for i, h in enumerate(norm_headers)
                 if any(c and (c in h or h in c) for c in cands)),
                -1,
            )
        if idx != -1:
            out[field] = idx
    return out


def _sheet_rows(wb, sheet_name: str) -> List[List[object]]:
    ws = wb[sheet_name]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _pick_sheet(wb, aliases: Dict[str, List[str]], prefer: re.Pattern) -> Optional[Tuple[str, int]]:
    best: Optional[Tuple[str, int]] = None
    for name in wb.sheetnames:
        rows = _sheet_rows(wb, name)
        if not rows or not rows[0]:
            continue
        headers = [str(h) if h is not None else "" for h in rows[0]]
        score = len(_build_column_map(headers, aliases))
        if prefer.search(name):
            score += 100
        if best is None or score > best[1]:
            best = (name, score)
    return best


def _rows_to_df(rows: List[List[object]], aliases: Dict[str, List[str]],
                numeric: set, date_fields: set) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    headers = [str(h) if h is not None else "" for h in rows[0]]
    col_map = _build_column_map(headers, aliases)
    out_rows: List[Dict[str, object]] = []
    for r in rows[1:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        row: Dict[str, object] = {}
        for field, idx in col_map.items():
            raw = r[idx] if idx < len(r) else None
            if field in date_fields:
                row[field] = _to_date(raw)
            elif field in numeric:
                row[field] = _to_num(raw)
            else:
                row[field] = "" if raw is None else str(raw)
        for field in aliases:
            row.setdefault(field, 0.0 if field in numeric else "")
        out_rows.append(row)
    return pd.DataFrame(out_rows)


def parse_workbook(buf: bytes) -> Tuple[pd.DataFrame, pd.DataFrame]:
    wb = load_workbook(io.BytesIO(buf), data_only=True, read_only=True)

    deal_pick = _pick_sheet(wb, DEAL_ALIASES, re.compile(r"raw|deal|extract|take", re.I))
    store_pick = _pick_sheet(wb, STORE_ALIASES, re.compile(r"store|active|merchant", re.I))

    deals = pd.DataFrame()
    # `prefer` bonus is +100; a real sheet match needs at least one alias hit too.
    if deal_pick and (deal_pick[1] % 100) > 0:
        deals = _rows_to_df(_sheet_rows(wb, deal_pick[0]), DEAL_ALIASES,
                            DEAL_NUMERIC, DEAL_DATE_FIELDS)

    stores = pd.DataFrame()
    if store_pick and (store_pick[1] % 100) > 0 and (not deal_pick or store_pick[0] != deal_pick[0]):
        stores = _rows_to_df(_sheet_rows(wb, store_pick[0]), STORE_ALIASES,
                             STORE_NUMERIC, STORE_DATE_FIELDS)
        if not stores.empty:
            stores.loc[stores["store_level"] == "", "store_level"] = "terr"

    if deals.empty:
        raise ValueError(
            "No recognizable deal rows found. Check that the raw extract tab has "
            "headers like trips, basket, gross_bookings."
        )

    return normalize_deals(deals), normalize_stores(stores)


def parse_excel_file(file_like) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if hasattr(file_like, "read"):
        buf = file_like.read()
    else:
        with open(file_like, "rb") as f:
            buf = f.read()
    return parse_workbook(buf)
